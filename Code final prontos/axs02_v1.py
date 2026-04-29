# -*- coding: utf-8 -*-
"""
AXS Energia Unidade 02 - CRI 46EUS e Debenture AXSD11 - IPCA + 11% a.a.

Objetivo
- Gerar historico diario de PU e fluxo de eventos para as duas pontas da
  operacao AXS 02: CRI e Debenture.
- Manter uma saida parecida com a usada no calculo da AXS 10, com CSV e XLSX.
- Reproduzir a metodologia observada no historico de PUs da Vortx da debenture:
  IPCA interpolado por dias uteis, juros 11% a.a. em base 252, fator de juros
  com 9 casas e valores monetarios com 8 casas sem arredondamento.

Ponto importante da emissao
- A Debenture usa NIk = mes imediatamente anterior ao pagamento.
- O CRI usa NIk = segundo mes imediatamente anterior ao pagamento.

Arquivos gerados
    controle_divida_axs02_v1_eventos.csv
    controle_divida_axs02_v1_eventos.xlsx
    historico_pu_axs02_v1_diario.csv
    historico_pu_axs02_v1_diario.xlsx
    validacao_vortx_axs02_deb.csv / .xlsx, se o historico AXSD11 for achado
"""

from __future__ import annotations

import csv
import json
import math
import ssl
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_DOWN, ROUND_HALF_UP, getcontext
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from urllib.request import Request, urlopen

getcontext().prec = 40

BASE_DIR = Path(__file__).resolve().parent
ROOT_DOCS_AXS02 = Path(r"C:\Users\rodol\AXS ENERGIA S A\Financeiro - Documentos\zDCM\AXS 02")

DATA_INTEGRALIZACAO = date(2022, 12, 29)
DATA_VENCIMENTO = date(2036, 12, 15)
PU_INICIAL = Decimal("1000.00000000")
TAXA_JUROS_AA = Decimal("0.11000000")
BASE_DU = Decimal("252")
DUT_PRIMEIRO_PERIODO = 22
IPCA_FUTURO_MENSAL = Decimal("0.0045000000")

ARQ_EVENTOS_CSV = BASE_DIR / "controle_divida_axs02_v1_eventos.csv"
ARQ_EVENTOS_XLSX = BASE_DIR / "controle_divida_axs02_v1_eventos.xlsx"
ARQ_DIARIO_CSV = BASE_DIR / "historico_pu_axs02_v1_diario.csv"
ARQ_DIARIO_XLSX = BASE_DIR / "historico_pu_axs02_v1_diario.xlsx"
ARQ_VALIDACAO_CSV = BASE_DIR / "validacao_vortx_axs02_deb.csv"
ARQ_VALIDACAO_XLSX = BASE_DIR / "validacao_vortx_axs02_deb.xlsx"


@dataclass(frozen=True)
class Instrumento:
    nome: str
    tipo: str
    codigo_if: str
    isin: str
    quantidade: Decimal
    lag_ipca_meses: int


INSTRUMENTOS = [
    Instrumento(
        nome="AXS II - Emissao 46 / Serie Unica",
        tipo="CRI",
        codigo_if="22L1467623",
        isin="BRRBRACRIG06",
        quantidade=Decimal("45000"),
        lag_ipca_meses=2,
    ),
    Instrumento(
        nome="AXS 02 - Emissao 1 / Serie Unica",
        tipo="DEBENTURE",
        codigo_if="AXSD11",
        isin="BRAXSDDBS005",
        quantidade=Decimal("40000"),
        lag_ipca_meses=1,
    ),
]

# Cronograma do anexo: data de pagamento e percentual do saldo de VNA amortizado.
CRONOGRAMA_RAW = [
    ("2023-01-16", "0.0000"), ("2023-02-15", "0.0000"), ("2023-03-15", "0.0000"),
    ("2023-04-17", "0.0000"), ("2023-05-15", "0.0000"), ("2023-06-15", "0.0000"),
    ("2023-07-17", "0.0000"), ("2023-08-15", "0.0000"), ("2023-09-15", "0.0000"),
    ("2023-10-16", "0.0000"), ("2023-11-16", "0.0000"), ("2023-12-15", "0.2000"),
    ("2024-01-15", "0.1002"), ("2024-02-15", "0.1003"), ("2024-03-15", "0.1004"),
    ("2024-04-15", "0.1005"), ("2024-05-15", "0.1006"), ("2024-06-17", "0.1007"),
    ("2024-07-15", "0.1008"), ("2024-08-15", "0.1009"), ("2024-09-16", "0.1010"),
    ("2024-10-15", "0.1011"), ("2024-11-18", "0.1012"), ("2024-12-16", "0.1013"),
    ("2025-01-15", "0.3550"), ("2025-02-17", "0.3562"), ("2025-03-17", "0.3575"),
    ("2025-04-15", "0.3588"), ("2025-05-15", "0.3601"), ("2025-06-16", "0.3614"),
    ("2025-07-15", "0.3627"), ("2025-08-15", "0.3640"), ("2025-09-15", "0.3653"),
    ("2025-10-15", "0.3667"), ("2025-11-17", "0.3680"), ("2025-12-15", "0.4749"),
    ("2026-01-15", "0.4242"), ("2026-02-18", "0.4260"), ("2026-03-16", "0.4278"),
    ("2026-04-15", "0.4296"), ("2026-05-15", "0.4315"), ("2026-06-15", "0.4334"),
    ("2026-07-15", "0.4353"), ("2026-08-17", "0.4372"), ("2026-09-15", "0.4391"),
    ("2026-10-15", "0.4410"), ("2026-11-16", "0.4430"), ("2026-12-15", "0.6674"),
    ("2027-01-15", "0.6719"), ("2027-02-15", "0.6764"), ("2027-03-15", "0.6810"),
    ("2027-04-15", "0.6857"), ("2027-05-17", "0.6904"), ("2027-06-15", "0.6952"),
    ("2027-07-15", "0.7001"), ("2027-08-16", "0.7051"), ("2027-09-15", "0.7101"),
    ("2027-10-15", "0.7151"), ("2027-11-16", "0.7203"), ("2027-12-15", "0.7255"),
    ("2028-01-17", "0.7308"), ("2028-02-15", "0.7362"), ("2028-03-15", "0.7417"),
    ("2028-04-17", "0.7472"), ("2028-05-15", "0.7528"), ("2028-06-16", "0.7585"),
    ("2028-07-17", "0.7643"), ("2028-08-15", "0.7702"), ("2028-09-15", "0.7762"),
    ("2028-10-16", "0.7823"), ("2028-11-16", "0.7884"), ("2028-12-15", "0.7947"),
    ("2029-01-15", "0.8011"), ("2029-02-15", "0.8075"), ("2029-03-15", "0.8141"),
    ("2029-04-16", "0.8208"), ("2029-05-15", "0.8276"), ("2029-06-15", "0.8345"),
    ("2029-07-16", "0.8415"), ("2029-08-15", "0.8487"), ("2029-09-17", "0.8559"),
    ("2029-10-15", "0.8633"), ("2029-11-16", "0.8708"), ("2029-12-17", "0.8785"),
    ("2030-01-15", "0.8863"), ("2030-02-15", "0.8942"), ("2030-03-15", "0.9023"),
    ("2030-04-15", "1.0622"), ("2030-05-15", "0.9202"), ("2030-06-17", "0.9288"),
    ("2030-07-15", "0.9375"), ("2030-08-15", "1.1041"), ("2030-09-16", "1.1164"),
    ("2030-10-15", "1.2903"), ("2030-11-18", "1.1438"), ("2030-12-16", "1.3223"),
    ("2031-01-15", "1.1725"), ("2031-02-17", "1.0169"), ("2031-03-17", "1.0274"),
    ("2031-04-15", "1.2111"), ("2031-05-15", "1.2259"), ("2031-06-16", "1.2411"),
    ("2031-07-15", "1.4363"), ("2031-08-15", "1.8215"), ("2031-09-15", "1.8553"),
    ("2031-10-15", "1.8904"), ("2031-11-17", "1.9268"), ("2031-12-15", "1.3752"),
    ("2032-01-15", "1.3944"), ("2032-02-16", "1.6162"), ("2032-03-15", "1.4374"),
    ("2032-04-15", "1.6667"), ("2032-05-17", "1.8008"), ("2032-06-15", "1.8339"),
    ("2032-07-15", "1.8681"), ("2032-08-16", "1.3438"), ("2032-09-15", "1.9296"),
    ("2032-10-15", "1.3889"), ("2032-11-16", "1.4085"), ("2032-12-15", "2.6190"),
    ("2033-01-17", "2.2005"), ("2033-02-15", "2.2500"), ("2033-03-15", "2.3018"),
    ("2033-04-18", "2.6178"), ("2033-05-16", "2.4194"), ("2033-06-15", "2.4793"),
    ("2033-07-15", "1.9774"), ("2033-08-15", "2.0173"), ("2033-09-15", "2.0588"),
    ("2033-10-17", "2.1021"), ("2033-11-16", "2.1472"), ("2033-12-15", "2.5078"),
    ("2034-01-16", "1.6077"), ("2034-02-15", "1.6340"), ("2034-03-15", "1.6611"),
    ("2034-04-17", "1.6892"), ("2034-05-15", "1.7182"), ("2034-06-15", "1.7483"),
    ("2034-07-17", "1.7794"), ("2034-08-15", "3.2609"), ("2034-09-15", "3.3708"),
    ("2034-10-16", "3.4884"), ("2034-11-16", "3.6145"), ("2034-12-15", "3.7500"),
    ("2035-01-15", "3.8961"), ("2035-02-15", "4.0541"), ("2035-03-15", "4.2254"),
    ("2035-04-16", "4.4118"), ("2035-05-15", "4.6154"), ("2035-06-15", "4.8387"),
    ("2035-07-16", "5.0847"), ("2035-08-15", "5.3571"), ("2035-09-17", "5.6604"),
    ("2035-10-15", "6.0000"), ("2035-11-16", "7.0922"), ("2035-12-17", "6.8702"),
    ("2036-01-15", "8.1967"), ("2036-02-15", "8.9286"), ("2036-03-17", "9.8039"),
    ("2036-04-15", "10.8696"), ("2036-05-15", "12.1951"), ("2036-06-16", "13.8889"),
    ("2036-07-15", "16.1290"), ("2036-08-15", "19.2308"), ("2036-09-15", "23.8095"),
    ("2036-10-15", "31.2500"), ("2036-11-17", "45.4545"), ("2036-12-15", "100.0000"),
]

CRONOGRAMA = [
    (datetime.strptime(data_txt, "%Y-%m-%d").date(), Decimal(perc_txt) / Decimal("100"))
    for data_txt, perc_txt in CRONOGRAMA_RAW
]

IPCA_INDICE_FALLBACK = {
    "202210": Decimal("6407.9300000000000"),
    "202211": Decimal("6434.2000000000000"),
    "202212": Decimal("6474.0900000000000"),
    "202301": Decimal("6508.4000000000000"),
    "202302": Decimal("6563.0700000000000"),
}


def trunc_dec(valor: Decimal, casas: int = 8) -> Decimal:
    return valor.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_DOWN)


def round_dec(valor: Decimal, casas: int = 2) -> Decimal:
    return valor.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_HALF_UP)


def trunc_float_dec(valor: float | Decimal, casas: int = 8) -> Decimal:
    escala = 10 ** casas
    truncado = math.floor(float(valor) * escala) / escala
    return Decimal(str(truncado)).quantize(Decimal("1").scaleb(-casas), rounding=ROUND_DOWN)


def data_ptbr(dt: date) -> str:
    return dt.strftime("%d/%m/%Y")


def periodo_yyyymm(dt: date) -> str:
    return f"{dt.year:04d}{dt.month:02d}"


def add_months_periodo(periodo: str, meses: int) -> str:
    ano = int(periodo[:4])
    mes = int(periodo[4:])
    total = ano * 12 + (mes - 1) + meses
    novo_ano = total // 12
    novo_mes = total % 12 + 1
    return f"{novo_ano:04d}{novo_mes:02d}"


def add_months_data(dt: date, meses: int) -> date:
    return datetime.strptime(add_months_periodo(periodo_yyyymm(dt), meses), "%Y%m").date()


def iter_periodos(inicio: str, fim: str) -> Iterable[str]:
    atual = inicio
    while atual <= fim:
        yield atual
        atual = add_months_periodo(atual, 1)


def easter_date(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def feriados_nacionais(start_year: int = 2022, end_year: int = 2036) -> set[date]:
    feriados: set[date] = set()
    for ano in range(start_year, end_year + 1):
        pascoa = easter_date(ano)
        feriados.update({
            date(ano, 1, 1),
            pascoa - timedelta(days=48),
            pascoa - timedelta(days=47),
            pascoa - timedelta(days=2),
            date(ano, 4, 21),
            date(ano, 5, 1),
            pascoa + timedelta(days=60),
            date(ano, 9, 7),
            date(ano, 10, 12),
            date(ano, 11, 2),
            date(ano, 11, 15),
            date(ano, 12, 25),
        })
        if ano >= 2024:
            feriados.add(date(ano, 11, 20))
    return feriados


FERIADOS = feriados_nacionais()


def eh_dia_util(dt: date) -> bool:
    return dt.weekday() < 5 and dt not in FERIADOS


def iter_dias_uteis_periodo(inicio: date, fim: date) -> Iterable[date]:
    atual = inicio + timedelta(days=1)
    while atual <= fim:
        if eh_dia_util(atual):
            yield atual
        atual += timedelta(days=1)


def contar_dias_uteis(inicio: date, fim: date) -> int:
    return sum(1 for _ in iter_dias_uteis_periodo(inicio, fim))


def obter_json_url(url: str, timeout: int = 60) -> object:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    ctx = ssl.create_default_context()
    with urlopen(req, timeout=timeout, context=ctx) as resp:
        return json.loads(resp.read().decode("utf-8"))


def obter_ipca_sidra(periodo_inicial: str, periodo_final: str) -> Tuple[Dict[str, Decimal], str]:
    url = (
        "https://apisidra.ibge.gov.br/values/"
        f"t/1737/n1/all/v/2266/p/{periodo_inicial}-{periodo_final}?formato=json"
    )
    try:
        dados = obter_json_url(url)
    except Exception as exc:
        indices = dict(IPCA_INDICE_FALLBACK)
        return indices, f"Fallback local IPCA inicial; falha SIDRA: {exc}"

    if not isinstance(dados, list) or len(dados) <= 1:
        indices = dict(IPCA_INDICE_FALLBACK)
        return indices, f"Fallback local IPCA inicial; SIDRA retornou vazio | {url}"

    indices: Dict[str, Decimal] = {}
    for item in dados[1:]:
        if not isinstance(item, dict):
            continue
        periodo = str(item.get("D3C", ""))
        valor = str(item.get("V", "")).replace(",", ".")
        if periodo and valor not in {"", "..."}:
            indices[periodo] = Decimal(valor)

    if not indices:
        indices = dict(IPCA_INDICE_FALLBACK)
        return indices, f"Fallback local IPCA inicial; SIDRA sem valores validos | {url}"

    return indices, f"IBGE SIDRA tabela 1737 variavel 2266 | {url}"


def preparar_indices_ipca(periodo_inicial: str, periodo_final: str) -> Tuple[Dict[str, Decimal], Dict[str, str], str]:
    indices, fonte = obter_ipca_sidra(periodo_inicial, periodo_final)
    fontes = {periodo: "IBGE/SIDRA numero-indice IPCA" for periodo in indices}

    if periodo_inicial not in indices:
        for periodo, valor in IPCA_INDICE_FALLBACK.items():
            indices.setdefault(periodo, valor)
            fontes.setdefault(periodo, "Fallback local IPCA inicial")

    periodos_disponiveis = [p for p in indices if periodo_inicial <= p <= periodo_final]
    if not periodos_disponiveis:
        raise RuntimeError("Nao foi possivel carregar a serie de IPCA necessaria.")

    ultimo_periodo = max(periodos_disponiveis)
    ultimo_indice = indices[ultimo_periodo]
    for periodo in iter_periodos(add_months_periodo(ultimo_periodo, 1), periodo_final):
        ultimo_indice = trunc_dec(ultimo_indice * (Decimal("1") + IPCA_FUTURO_MENSAL), 13)
        indices[periodo] = ultimo_indice
        fontes[periodo] = f"Projetado fixo {IPCA_FUTURO_MENSAL * Decimal('100')}% a.m."

    return indices, fontes, fonte


def meses_ipca(data_pagto: date, lag_meses: int) -> Tuple[str, str]:
    primeiro_dia_mes = date(data_pagto.year, data_pagto.month, 1)
    nik = periodo_yyyymm(add_months_data(primeiro_dia_mes, -lag_meses))
    nik_anterior = add_months_periodo(nik, -1)
    return nik, nik_anterior


def fator_ipca(
    indices_ipca: Dict[str, Decimal],
    data_pagto: date,
    lag_meses: int,
    dup: int,
    dut: int,
) -> Tuple[Decimal, str, str, Decimal, Decimal]:
    nik, nik_anterior = meses_ipca(data_pagto, lag_meses)
    if nik not in indices_ipca or nik_anterior not in indices_ipca:
        raise RuntimeError(f"IPCA ausente para {nik}/{nik_anterior}.")

    if dup == 0:
        return Decimal("1.00000000"), nik, nik_anterior, indices_ipca[nik], indices_ipca[nik_anterior]

    # A Vortx trunca a razao mensal antes da interpolacao por DU.
    razao = float(trunc_dec(indices_ipca[nik] / indices_ipca[nik_anterior], 8))
    fator = math.pow(razao, dup / dut)
    return trunc_float_dec(fator, 8), nik, nik_anterior, indices_ipca[nik], indices_ipca[nik_anterior]


def fator_juros(du: int) -> Decimal:
    if du == 0:
        return Decimal("1.000000000")
    bruto = (Decimal("1") + TAXA_JUROS_AA) ** (Decimal(du) / BASE_DU)
    return round_dec(bruto, 9)


def caminho_alternativo(caminho: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return caminho.with_name(f"{caminho.stem}_{timestamp}{caminho.suffix}")


def salvar_com_fallback(funcao_salvar, caminho: Path, *args) -> Tuple[Path, bool]:
    try:
        funcao_salvar(*args, caminho)
        return caminho, False
    except PermissionError:
        alternativo = caminho_alternativo(caminho)
        funcao_salvar(*args, alternativo)
        return alternativo, True


def decimal_para_csv(valor: object) -> object:
    if isinstance(valor, Decimal):
        return format(valor, "f").replace(".", ",")
    return valor


def decimal_para_excel(valor: object) -> object:
    if isinstance(valor, Decimal):
        return float(valor)
    return valor


def salvar_csv(linhas: List[Dict[str, object]], caminho: Path) -> None:
    if not linhas:
        raise RuntimeError(f"Nenhuma linha para salvar em {caminho.name}.")
    caminho.parent.mkdir(parents=True, exist_ok=True)
    campos = list(linhas[0].keys())
    with caminho.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos, delimiter=";")
        writer.writeheader()
        for linha in linhas:
            writer.writerow({k: decimal_para_csv(v) for k, v in linha.items()})


def salvar_xlsx(linhas: List[Dict[str, object]], caminho: Path) -> None:
    if not linhas:
        raise RuntimeError(f"Nenhuma linha para salvar em {caminho.name}.")
    import pandas as pd

    caminho.parent.mkdir(parents=True, exist_ok=True)
    dados = [{k: decimal_para_excel(v) for k, v in linha.items()} for linha in linhas]
    pd.DataFrame(dados).to_excel(caminho, index=False)


def salvar_workbook(
    eventos: List[Dict[str, object]],
    diario: List[Dict[str, object]],
    parametros: List[Dict[str, object]],
    validacao: List[Dict[str, object]],
    caminho: Path,
) -> None:
    import pandas as pd

    caminho.parent.mkdir(parents=True, exist_ok=True)

    def to_df(linhas: List[Dict[str, object]]) -> "pd.DataFrame":
        return pd.DataFrame([{k: decimal_para_excel(v) for k, v in linha.items()} for linha in linhas])

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        to_df(eventos).to_excel(writer, sheet_name="Eventos", index=False)
        to_df(diario).to_excel(writer, sheet_name="Historico_PU", index=False)
        to_df(parametros).to_excel(writer, sheet_name="Parametros", index=False)
        if validacao:
            to_df(validacao).to_excel(writer, sheet_name="Validacao_DEB_Vortx", index=False)


def datas_diarias_periodo(inicio: date, fim: date, primeiro_periodo: bool) -> Iterable[date]:
    if primeiro_periodo and eh_dia_util(inicio):
        yield inicio
    yield from iter_dias_uteis_periodo(inicio, fim)


def calcular_instrumento(
    instrumento: Instrumento,
    indices_ipca: Dict[str, Decimal],
    fontes_ipca: Dict[str, str],
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    eventos: List[Dict[str, object]] = []
    diario: List[Dict[str, object]] = []

    saldo_base = PU_INICIAL
    inicio_periodo = DATA_INTEGRALIZACAO

    for idx, (data_pagto, perc_amort) in enumerate(CRONOGRAMA):
        primeiro_periodo = idx == 0
        dut = DUT_PRIMEIRO_PERIODO if primeiro_periodo else contar_dias_uteis(inicio_periodo, data_pagto)
        if dut <= 0:
            raise RuntimeError(f"DUT invalido no periodo ate {data_ptbr(data_pagto)}.")

        evento_linha: Dict[str, object] | None = None
        for data_calc in datas_diarias_periodo(inicio_periodo, data_pagto, primeiro_periodo):
            dup = 0 if data_calc == inicio_periodo else contar_dias_uteis(inicio_periodo, data_calc)
            fator_c, nik, nik_anterior, valor_nik, valor_nik_anterior = fator_ipca(
                indices_ipca, data_pagto, instrumento.lag_ipca_meses, dup, dut
            )
            fator_j = fator_juros(dup)
            vna = trunc_float_dec(float(saldo_base) * float(fator_c), 8)
            juros = trunc_dec(vna * (fator_j - Decimal("1")), 8) if dup else Decimal("0.00000000")

            linha_evento = data_calc == data_pagto
            amort = trunc_dec(vna * perc_amort, 8) if linha_evento else Decimal("0.00000000")
            if linha_evento and idx == len(CRONOGRAMA) - 1:
                amort = vna

            pu_cheio = trunc_dec(vna + juros, 8)
            pu_vazio = trunc_dec(vna - amort, 8) if linha_evento else pu_cheio
            total = trunc_dec(juros + amort, 8) if linha_evento else Decimal("0.00000000")
            saldo_evento = trunc_dec(vna - amort, 8) if linha_evento else Decimal("0.00000000")
            fonte_periodo = " | ".join(sorted({
                fontes_ipca.get(nik, ""),
                fontes_ipca.get(nik_anterior, ""),
            })).strip(" |")

            diario.append({
                "Instrumento": instrumento.tipo,
                "Codigo_IF": instrumento.codigo_if,
                "ISIN": instrumento.isin,
                "Data": data_ptbr(data_calc),
                "Data_ISO": data_calc.isoformat(),
                "Data_Inicio_Periodo": data_ptbr(inicio_periodo),
                "Data_Proximo_Evento": data_ptbr(data_pagto),
                "DU_Periodo": dup,
                "DUT_IPCA": dut,
                "Mes_NIk": nik,
                "Mes_NIk_1": nik_anterior,
                "NIk": valor_nik,
                "NIk_1": valor_nik_anterior,
                "Fator_C_IPCA": fator_c,
                "Fator_Juros": fator_j,
                "Valor_Nominal": vna,
                "Valor_dos_Juros": juros,
                "PU_Cheio": pu_cheio,
                "PU_Vazio": pu_vazio,
                "Juros_%": round_dec(juros, 2) if linha_evento else Decimal("0.00"),
                "Amortizacao": amort,
                "Total": total,
                "Saldo_Pos_Evento": saldo_evento if linha_evento else "",
                "Fonte_IPCA": fonte_periodo,
            })

            if linha_evento:
                juros_rs = round_dec(juros * instrumento.quantidade, 2)
                amort_rs = round_dec(amort * instrumento.quantidade, 2)
                evento_linha = {
                    "Instrumento": instrumento.tipo,
                    "Nome": instrumento.nome,
                    "Codigo_IF": instrumento.codigo_if,
                    "ISIN": instrumento.isin,
                    "Data_Ref": data_ptbr(data_pagto),
                    "Data_Pgto": data_ptbr(data_pagto),
                    "Evento": "Juros + Amortizacao" if amort else "Juros",
                    "DU_Juros": dup,
                    "DU_IPCA": dup,
                    "DUT_IPCA": dut,
                    "Mes_NIk": nik,
                    "Mes_NIk_1": nik_anterior,
                    "NIk": valor_nik,
                    "NIk_1": valor_nik_anterior,
                    "Fator_C_IPCA": fator_c,
                    "Fator_Juros": fator_j,
                    "Perc_Amort": perc_amort,
                    "PU_VNa_Ini": saldo_base,
                    "PU_VNa_Atualizado": vna,
                    "PU_Juros": juros,
                    "PU_Amort": amort,
                    "PU_Total": total,
                    "PU_VNa_Fim": saldo_evento,
                    "Juros_R$": juros_rs,
                    "Amort_R$": amort_rs,
                    "PMT_Total": round_dec(juros_rs + amort_rs, 2),
                    "Saldo_Devedor_R$": round_dec(saldo_evento * instrumento.quantidade, 2),
                    "Fonte_IPCA": fonte_periodo,
                }

        if evento_linha is None:
            raise RuntimeError(f"Evento nao calculado para {instrumento.codigo_if} em {data_ptbr(data_pagto)}.")

        eventos.append(evento_linha)
        saldo_base = evento_linha["PU_VNa_Fim"]  # type: ignore[assignment]
        inicio_periodo = data_pagto

    return eventos, diario


def encontrar_historico_vortx_deb() -> Path | None:
    if not ROOT_DOCS_AXS02.exists():
        return None
    candidatos = list(ROOT_DOCS_AXS02.rglob("historico-pus-AXSD11.xlsx"))
    return candidatos[0] if candidatos else None


def carregar_historico_vortx_deb() -> Tuple[Path | None, Dict[date, Dict[str, Decimal]]]:
    caminho = encontrar_historico_vortx_deb()
    if caminho is None:
        return None, {}

    from openpyxl import load_workbook

    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    historico: Dict[date, Dict[str, Decimal]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_cell = row[0]
        if data_cell is None:
            continue
        data_hist = data_cell.date() if isinstance(data_cell, datetime) else data_cell
        historico[data_hist] = {
            "Valor_Nominal": Decimal(str(row[1] or 0)),
            "Valor_dos_Juros": Decimal(str(row[2] or 0)),
            "PU_Cheio": Decimal(str(row[3] or 0)),
            "PU_Vazio": Decimal(str(row[4] or 0)),
            "Amortizacao": Decimal(str(row[6] or 0)),
            "Total": Decimal(str(row[7] or 0)),
        }
    return caminho, historico


def data_publicacao_aproximada_ipca(periodo: str) -> date:
    ano = int(periodo[:4])
    mes = int(periodo[4:])
    proximo = add_months_data(date(ano, mes, 1), 1)
    return date(proximo.year, proximo.month, 11)


def montar_validacao_vortx(diario: List[Dict[str, object]]) -> Tuple[Path | None, List[Dict[str, object]]]:
    caminho, vortx = carregar_historico_vortx_deb()
    if not vortx:
        return caminho, []

    calc_deb: Dict[date, Dict[str, object]] = {}
    for linha in diario:
        if linha["Instrumento"] != "DEBENTURE":
            continue
        calc_deb[datetime.strptime(str(linha["Data"]), "%d/%m/%Y").date()] = linha

    validacao: List[Dict[str, object]] = []
    data_base_vortx = max(vortx)
    for data_hist, linha_vortx in sorted(vortx.items()):
        linha_calc = calc_deb.get(data_hist)
        if linha_calc is None:
            validacao.append({
                "Data": data_ptbr(data_hist),
                "Status": "Data ausente no calculo",
                "Vortx_PU_Cheio": linha_vortx["PU_Cheio"],
            })
            continue

        mes_nik = str(linha_calc["Mes_NIk"])
        comparavel = data_base_vortx >= data_publicacao_aproximada_ipca(mes_nik)
        row: Dict[str, object] = {
            "Data": data_ptbr(data_hist),
            "Status": "OK",
            "Mes_NIk": mes_nik,
            "IPCA_Comparavel": "SIM" if comparavel else "NAO",
            "Observacao": "" if comparavel else "Vortx provavelmente usava IPCA projetado nesta data",
        }
        for campo in ["Valor_Nominal", "Valor_dos_Juros", "PU_Cheio", "PU_Vazio", "Amortizacao", "Total"]:
            calc_val = linha_calc[campo]
            calc_dec = calc_val if isinstance(calc_val, Decimal) else Decimal(str(calc_val))
            vortx_val = linha_vortx[campo]
            row[f"Calc_{campo}"] = calc_dec
            row[f"Vortx_{campo}"] = vortx_val
            row[f"Dif_{campo}"] = trunc_dec(calc_dec - vortx_val, 8)
        validacao.append(row)

    return caminho, validacao


def parametros_saida(fonte_ipca: str, caminho_vortx: Path | None) -> List[Dict[str, object]]:
    linhas: List[Dict[str, object]] = []
    for instrumento in INSTRUMENTOS:
        linhas.append({
            "Instrumento": instrumento.tipo,
            "Codigo_IF": instrumento.codigo_if,
            "ISIN": instrumento.isin,
            "Quantidade": instrumento.quantidade,
            "PU_Inicial": PU_INICIAL,
            "Data_Integralizacao_Usada": data_ptbr(DATA_INTEGRALIZACAO),
            "Remuneracao": "IPCA + 11% a.a.",
            "Base_DU": BASE_DU,
            "Lag_IPCA_Meses": instrumento.lag_ipca_meses,
            "DUT_Primeiro_Periodo": DUT_PRIMEIRO_PERIODO,
            "IPCA_Futuro_Mensal": IPCA_FUTURO_MENSAL,
            "Fonte_IPCA": fonte_ipca,
            "Historico_Vortx_DEB": str(caminho_vortx) if caminho_vortx else "Nao encontrado",
        })
    return linhas


def max_abs_diferenca(validacao: List[Dict[str, object]], campo: str, somente_comparavel: bool = False) -> Decimal:
    valores = [
        abs(linha[f"Dif_{campo}"])
        for linha in validacao
        if f"Dif_{campo}" in linha and isinstance(linha[f"Dif_{campo}"], Decimal)
        and (not somente_comparavel or linha.get("IPCA_Comparavel") == "SIM")
    ]
    return max(valores) if valores else Decimal("0")


def main() -> None:
    menor_mes_ipca = "202210"
    maior_mes_ipca = max(meses_ipca(DATA_VENCIMENTO, inst.lag_ipca_meses)[0] for inst in INSTRUMENTOS)
    indices_ipca, fontes_ipca, fonte_ipca = preparar_indices_ipca(menor_mes_ipca, maior_mes_ipca)

    eventos: List[Dict[str, object]] = []
    diario: List[Dict[str, object]] = []
    for instrumento in INSTRUMENTOS:
        eventos_inst, diario_inst = calcular_instrumento(instrumento, indices_ipca, fontes_ipca)
        eventos.extend(eventos_inst)
        diario.extend(diario_inst)

    caminho_vortx, validacao = montar_validacao_vortx(diario)
    parametros = parametros_saida(fonte_ipca, caminho_vortx)

    saidas: List[Tuple[str, Path, bool]] = []
    for caminho, linhas in [
        (ARQ_EVENTOS_CSV, eventos),
        (ARQ_DIARIO_CSV, diario),
        (ARQ_VALIDACAO_CSV, validacao),
    ]:
        if linhas:
            gerado, fallback = salvar_com_fallback(salvar_csv, caminho, linhas)
            saidas.append((caminho.name, gerado, fallback))

    for caminho, linhas in [
        (ARQ_EVENTOS_XLSX, eventos),
        (ARQ_DIARIO_XLSX, diario),
        (ARQ_VALIDACAO_XLSX, validacao),
    ]:
        if linhas:
            gerado, fallback = salvar_com_fallback(salvar_xlsx, caminho, linhas)
            saidas.append((caminho.name, gerado, fallback))

    workbook_completo = BASE_DIR / "controle_divida_axs02_v1_completo.xlsx"
    gerado, fallback = salvar_com_fallback(
        salvar_workbook, workbook_completo, eventos, diario, parametros, validacao
    )
    saidas.append((workbook_completo.name, gerado, fallback))

    print("Fonte IPCA:", fonte_ipca)
    print("Arquivos gerados:")
    for nome_padrao, caminho, fallback in saidas:
        aviso = " (alternativo; padrao estava aberto)" if fallback else ""
        print(f"- {caminho}{aviso}")

    if validacao:
        ultima_data = validacao[-1]["Data"]
        comparaveis = [linha for linha in validacao if linha.get("IPCA_Comparavel") == "SIM"]
        print("\nValidacao Vortx DEB AXSD11:")
        print(f"- Arquivo: {caminho_vortx}")
        print(f"- Linhas comparadas: {len(validacao)} ate {ultima_data}")
        print(f"- Linhas com IPCA oficial comparavel: {len(comparaveis)}")
        print(
            "- Max abs dif PU_Cheio em linhas comparaveis: "
            f"{format(max_abs_diferenca(validacao, 'PU_Cheio', True), 'f')}"
        )
        print(
            "- Max abs dif Valor_Nominal em linhas comparaveis: "
            f"{format(max_abs_diferenca(validacao, 'Valor_Nominal', True), 'f')}"
        )
        print(f"- Max abs dif PU_Cheio total: {format(max_abs_diferenca(validacao, 'PU_Cheio'), 'f')}")
        print(f"- Max abs dif Total: {format(max_abs_diferenca(validacao, 'Total'), 'f')}")

    print("\nPrimeiras linhas DEB calculadas:")
    for linha in [x for x in diario if x["Instrumento"] == "DEBENTURE"][:4]:
        print(
            f"{linha['Data']} | VN {linha['Valor_Nominal']} | Juros {linha['Valor_dos_Juros']} | "
            f"PU {linha['PU_Cheio']}"
        )


if __name__ == "__main__":
    main()
