# -*- coding: utf-8 -*-
"""
AXS Energia Unidade 09 - Debenture AXS09 - calculo PMT padrao ANBIMA com projecao Focus/BCB
Versao v1: modelo baseado nos motores AXS 07 e AXS 08, com exportacao CSV e XLSX.

Objetivo
- Calcular o fluxo da debenture AXS 09 sem imputar valores calculados pela Vortx.
- Usa IPCA oficial do SIDRA/IBGE quando disponivel e Focus/BCB para meses futuros.
- Regras principais da documentacao da emissao e do 1o aditivo:
  * Data de emissao: 20/09/2024.
  * Data de inicio da rentabilidade: data da primeira integralizacao. Parametro configuravel abaixo.
  * Quantidade: 93.000 debentures.
  * PU inicial: R$ 1.000,00.
  * Remuneracao: IPCA + 10,98% a.a., base 252 dias uteis, conforme 1o aditivo.
  * Juros pagos semestralmente em marco e setembro.
  * Juros dos dois primeiros periodos de capitalizacao incorporados em 15/03/2025 e 15/09/2025.
  * Amortizacao programada em 26 parcelas semestrais, de 15/03/2026 ate 15/09/2038.

Como rodar
    python axs09_v1_focus_bcb_xlsx.py

Arquivos gerados
    controle_divida_axs09_v1_focus.csv
    controle_divida_axs09_v1_focus.xlsx

Dependencias opcionais
    pip install openpyxl python-bcb

Observacao
- O script calcula pelo metodo, nao usa PU, VNA, juros ou PMT da Vortx como entrada.
- Se a primeira integralizacao efetiva da AXS 09 tiver ocorrido em data diferente, ajuste
  DATA_INICIO_RENTABILIDADE abaixo e rode novamente.
"""

from __future__ import annotations

import csv
import json
import ssl
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_DOWN, ROUND_HALF_UP, getcontext
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.request import Request, urlopen
from urllib.parse import quote, urlencode
import importlib.util

getcontext().prec = 34

DATA_EMISSAO = date(2024, 9, 20)
# A escritura define a Data de Inicio da Rentabilidade como a data da primeira integralizacao.
# O 1o aditivo foi assinado em 03/10/2024 e informa que as debentures ainda nao haviam sido integralizadas.
# Ajuste esta data caso o extrato/historico de PUs da Vortx indique outro primeiro dia efetivo.
DATA_INICIO_RENTABILIDADE = date(2024, 10, 4)
DATAS_INCORPORACAO_JUROS = [date(2025, 3, 15), date(2025, 9, 15)]
TAXA_AA = Decimal("0.1098")
QUANTIDADE = Decimal("93000")
PU_INICIAL = Decimal("1000.00000000")

# Cronograma do Anexo VI da escritura AXS 09.
# Usamos a coluna "Saldo do Valor Nominal Unitario Atualizado a ser amortizado",
# isto e, a TAI aplicada sobre o saldo VNA atualizado existente em cada data.
CRONOGRAMA_RAW = [
    ("2026-03-15", "0.5000"), ("2026-09-15", "3.0151"),
    ("2027-03-15", "2.7461"), ("2027-09-15", "2.9302"),
    ("2028-03-15", "3.0187"), ("2028-09-15", "3.3956"),
    ("2029-03-15", "3.5149"), ("2029-09-15", "4.2502"),
    ("2030-03-15", "4.0583"), ("2030-09-15", "4.6266"),
    ("2031-03-15", "5.5440"), ("2031-09-15", "5.8694"),
    ("2032-03-15", "6.2354"), ("2032-09-15", "6.6500"),
    ("2033-03-15", "7.1238"), ("2033-09-15", "6.7114"),
    ("2034-03-15", "9.2497"), ("2034-09-15", "7.9275"),
    ("2035-03-15", "9.8401"), ("2035-09-15", "12.2783"),
    ("2036-03-15", "13.9969"), ("2036-09-15", "18.0832"),
    ("2037-03-15", "22.0751"), ("2037-09-15", "31.1615"),
    ("2038-03-15", "45.2675"), ("2038-09-15", "100.0000"),
]

CRONOGRAMA = [(datetime.strptime(d, "%Y-%m-%d").date(), Decimal(p) / Decimal("100")) for d, p in CRONOGRAMA_RAW]

# Fallback oficial/conhecido para a fase que estamos auditando. Se o SIDRA estiver disponivel,
# ele sera usado no lugar destes valores.
IPCA_FALLBACK_VARIACAO_MENSAL = {
    "2024-02": Decimal("0.0083"), "2024-03": Decimal("0.0016"), "2024-04": Decimal("0.0038"),
    "2024-05": Decimal("0.0046"), "2024-06": Decimal("0.0021"), "2024-07": Decimal("0.0038"),
    "2024-08": Decimal("-0.0002"), "2024-09": Decimal("0.0044"), "2024-10": Decimal("0.0056"),
    "2024-11": Decimal("0.0039"), "2024-12": Decimal("0.0052"), "2025-01": Decimal("0.0016"),
    "2025-02": Decimal("0.0131"), "2025-03": Decimal("0.0056"), "2025-04": Decimal("0.0043"),
    "2025-05": Decimal("0.0026"), "2025-06": Decimal("0.0024"), "2025-07": Decimal("0.0026"),
    "2025-08": Decimal("-0.0011"), "2025-09": Decimal("0.0048"), "2025-10": Decimal("0.0009"),
    "2025-11": Decimal("0.0018"), "2025-12": Decimal("0.0033"), "2026-01": Decimal("0.0033"),
    "2026-02": Decimal("0.0070"), "2026-03": Decimal("0.0088"),
}
IPCA_PROJECAO_MENSAL_PADRAO = Decimal("0.0045")


def trunc_dec(x: Decimal, casas: int = 8) -> Decimal:
    return x.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_DOWN)


def round_dec(x: Decimal, casas: int = 2) -> Decimal:
    return x.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_HALF_UP)


def add_months(dt: date, months: int) -> date:
    y = dt.year + (dt.month - 1 + months) // 12
    m = (dt.month - 1 + months) % 12 + 1
    return date(y, m, 1)


def mes_str(dt: date) -> str:
    return f"{dt.year:04d}-{dt.month:02d}"


def aniversario(data_pagto: date) -> date:
    return date(data_pagto.year, data_pagto.month, 15)


def easter_date(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def feriados_nacionais(start_year: int = 2024, end_year: int = 2041) -> set[date]:
    fs: set[date] = set()
    for y in range(start_year, end_year + 1):
        pascoa = easter_date(y)
        fs.update({
            date(y, 1, 1),
            pascoa - timedelta(days=48),  # Carnaval segunda
            pascoa - timedelta(days=47),  # Carnaval terca
            pascoa - timedelta(days=2),   # Sexta-feira santa
            date(y, 4, 21),
            date(y, 5, 1),
            pascoa + timedelta(days=60),  # Corpus Christi
            date(y, 9, 7),
            date(y, 10, 12),
            date(y, 11, 2),
            date(y, 11, 15),
            date(y, 11, 20),
            date(y, 12, 25),
        })
    return fs


FERIADOS = feriados_nacionais()


def eh_dia_util(dt: date) -> bool:
    return dt.weekday() < 5 and dt not in FERIADOS


def dias_uteis(inicio: date, fim: date) -> int:
    """Conta dias uteis em [inicio, fim), conforme convencao de juros."""
    count = 0
    dt = inicio
    while dt < fim:
        if eh_dia_util(dt):
            count += 1
        dt += timedelta(days=1)
    return count


def periodo_final_sidra() -> str:
    max_aniv = max(aniversario(d) for d, _ in CRONOGRAMA)
    fim = add_months(max_aniv, 1)
    return f"{fim.year:04d}{fim.month:02d}"


def indices_fallback() -> Dict[str, Decimal]:
    idx = Decimal("7000.0000000000000")
    indices: Dict[str, Decimal] = {}
    for m in sorted(IPCA_FALLBACK_VARIACAO_MENSAL):
        idx = idx * (Decimal("1") + IPCA_FALLBACK_VARIACAO_MENSAL[m])
        indices[m] = idx
    return indices


def obter_ipca_numero_indice_sidra() -> Tuple[Dict[str, Decimal], str]:
    """Busca numero-indice IPCA no SIDRA. Se falhar, usa fallback local."""
    fim = periodo_final_sidra()
    url = f"https://apisidra.ibge.gov.br/values/t/1737/n1/all/v/2266/p/202402-{fim}?formato=json"
    try:
        ctx = ssl.create_default_context()
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=12, context=ctx) as resp:
            dados = json.loads(resp.read().decode("utf-8"))
        indices: Dict[str, Decimal] = {}
        for item in dados[1:]:
            periodo = str(item.get("D3C", ""))
            valor = str(item.get("V", "")).replace(",", ".")
            if len(periodo) == 6 and valor not in ("", "...", "-"):
                indices[f"{periodo[:4]}-{periodo[4:]}"] = Decimal(valor)
        if not indices:
            raise RuntimeError("SIDRA retornou vazio ou layout inesperado")
        return indices, f"SIDRA/IBGE Tabela 1737 v/2266 | {url}"
    except Exception as exc:
        return indices_fallback(), f"FALLBACK local por variacao mensal IPCA; motivo: {exc}"




def decimal_ptbr(valor: object) -> Decimal | None:
    if valor is None:
        return None
    txt = str(valor).strip().replace("%", "").replace(" ", "")
    if not txt or txt in ("...", "-", "--"):
        return None
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    else:
        txt = txt.replace(",", ".")
    try:
        return Decimal(txt)
    except Exception:
        return None


def parse_mes_referencia(valor: object) -> str | None:
    """Converte DataReferencia do Focus para AAAA-MM."""
    if valor is None:
        return None
    txt = str(valor).strip()
    if not txt:
        return None
    if len(txt) >= 7 and txt[4] == "-" and txt[:4].isdigit() and txt[5:7].isdigit():
        return txt[:7]
    if "/" in txt:
        partes = txt.split("/")
        if len(partes) >= 2 and partes[0].isdigit() and partes[1].isdigit():
            mes = int(partes[0])
            ano = int(partes[1])
            if ano < 100:
                ano += 2000
            if 1 <= mes <= 12:
                return f"{ano:04d}-{mes:02d}"
    if len(txt) == 6 and txt.isdigit():
        return f"{txt[:4]}-{txt[4:]}"
    return None


def obter_json_url(url: str, timeout: int = 20) -> object:
    ctx = ssl.create_default_context()
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=timeout, context=ctx) as resp:
        return json.loads(resp.read().decode("utf-8"))


def build_odata_url(recurso: str, params: Dict[str, str], usar_parenteses: bool = True) -> str:
    base = "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata"
    sufixo = "()" if usar_parenteses and not recurso.endswith(")") else ""
    return f"{base}/{recurso}{sufixo}?{urlencode(params, quote_via=quote)}"


def obter_focus_odata_mensal() -> Tuple[Dict[str, Decimal], str]:
    recursos = ["ExpectativaMercadoMensais", "ExpectativasMercadoMensais"]
    params = {
        "$top": "10000",
        "$format": "json",
        "$select": "Indicador,Data,DataReferencia,Mediana",
        "$filter": "Indicador eq 'IPCA'",
        "$orderby": "Data desc",
    }
    erros: List[str] = []
    for recurso in recursos:
        for usar_parenteses in (True, False):
            try:
                url = build_odata_url(recurso, params, usar_parenteses=usar_parenteses)
                dados = obter_json_url(url, timeout=30)
                itens = dados.get("value", []) if isinstance(dados, dict) else []
                out: Dict[str, Tuple[str, Decimal]] = {}
                for item in itens:
                    if str(item.get("Indicador", "")).upper() != "IPCA":
                        continue
                    mes = parse_mes_referencia(item.get("DataReferencia"))
                    med = decimal_ptbr(item.get("Mediana"))
                    if not mes or med is None:
                        continue
                    taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                    data_pub = str(item.get("Data", ""))
                    if mes not in out or data_pub > out[mes][0]:
                        out[mes] = (data_pub, taxa)
                if out:
                    nome = recurso + ("()" if usar_parenteses else "")
                    return {m: v[1] for m, v in out.items()}, f"Focus/BCB mensal via OData {nome}"
                erros.append(f"{recurso}: sem registros")
            except Exception as exc:
                erros.append(f"{recurso}: {exc}")
    return {}, "Focus/BCB mensal OData indisponivel: " + " | ".join(erros[:4])


def obter_focus_odata_anual() -> Tuple[Dict[int, Decimal], str]:
    recursos = ["ExpectativasMercadoAnuais", "ExpectativaMercadoAnuais"]
    params = {
        "$top": "10000",
        "$format": "json",
        "$select": "Indicador,Data,DataReferencia,Mediana",
        "$filter": "Indicador eq 'IPCA'",
        "$orderby": "Data desc",
    }
    erros: List[str] = []
    for recurso in recursos:
        for usar_parenteses in (True, False):
            try:
                url = build_odata_url(recurso, params, usar_parenteses=usar_parenteses)
                dados = obter_json_url(url, timeout=30)
                itens = dados.get("value", []) if isinstance(dados, dict) else []
                out: Dict[int, Tuple[str, Decimal]] = {}
                for item in itens:
                    if str(item.get("Indicador", "")).upper() != "IPCA":
                        continue
                    ref = str(item.get("DataReferencia", "")).strip()
                    if not ref.isdigit():
                        continue
                    ano = int(ref)
                    med = decimal_ptbr(item.get("Mediana"))
                    if med is None:
                        continue
                    taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                    data_pub = str(item.get("Data", ""))
                    if ano not in out or data_pub > out[ano][0]:
                        out[ano] = (data_pub, taxa)
                if out:
                    nome = recurso + ("()" if usar_parenteses else "")
                    return {a: v[1] for a, v in out.items()}, f"Focus/BCB anual via OData {nome}"
                erros.append(f"{recurso}: sem registros")
            except Exception as exc:
                erros.append(f"{recurso}: {exc}")
    return {}, "Focus/BCB anual OData indisponivel: " + " | ".join(erros[:4])


def obter_focus_python_bcb() -> Tuple[Dict[str, Decimal], Dict[int, Decimal], str]:
    """Tenta usar a biblioteca python-bcb. Se nao estiver instalada, retorna vazio."""
    try:
        if importlib.util.find_spec("bcb") is None:
            return {}, {}, "python-bcb nao instalado"
        from bcb import Expectativas  # type: ignore
        em = Expectativas()
    except Exception as exc:
        return {}, {}, f"python-bcb indisponivel: {exc}"

    mensagens: List[str] = []
    mensal: Dict[str, Decimal] = {}
    anual: Dict[int, Decimal] = {}

    try:
        ep = em.get_endpoint("ExpectativaMercadoMensais")
        df = (
            ep.query()
            .filter(ep.Indicador == "IPCA")
            .select(ep.Indicador, ep.Data, ep.DataReferencia, ep.Mediana)
            .orderby(ep.Data.desc())
            .limit(20000)
            .collect()
        )
        temp: Dict[str, Tuple[str, Decimal]] = {}
        for item in df.to_dict("records"):
            mes = parse_mes_referencia(item.get("DataReferencia"))
            med = decimal_ptbr(item.get("Mediana"))
            if not mes or med is None:
                continue
            taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
            data_pub = str(item.get("Data", ""))
            if mes not in temp or data_pub > temp[mes][0]:
                temp[mes] = (data_pub, taxa)
        mensal = {m: v[1] for m, v in temp.items()}
        mensagens.append(f"python-bcb mensal: {len(mensal)} meses")
    except Exception as exc:
        mensagens.append(f"python-bcb mensal falhou: {exc}")

    try:
        ep = em.get_endpoint("ExpectativasMercadoAnuais")
        df = (
            ep.query()
            .filter(ep.Indicador == "IPCA")
            .select(ep.Indicador, ep.Data, ep.DataReferencia, ep.Mediana)
            .orderby(ep.Data.desc())
            .limit(20000)
            .collect()
        )
        temp2: Dict[int, Tuple[str, Decimal]] = {}
        for item in df.to_dict("records"):
            ref = str(item.get("DataReferencia", "")).strip()
            med = decimal_ptbr(item.get("Mediana"))
            if not ref.isdigit() or med is None:
                continue
            ano = int(ref)
            taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
            data_pub = str(item.get("Data", ""))
            if ano not in temp2 or data_pub > temp2[ano][0]:
                temp2[ano] = (data_pub, taxa)
        anual = {a: v[1] for a, v in temp2.items()}
        mensagens.append(f"python-bcb anual: {len(anual)} anos")
    except Exception as exc:
        mensagens.append(f"python-bcb anual falhou: {exc}")

    return mensal, anual, " | ".join(mensagens)


def obter_focus_ipca() -> Tuple[Dict[str, Decimal], Dict[int, Decimal], str]:
    mensal, anual, fonte_py = obter_focus_python_bcb()
    fontes = [fonte_py]
    if not mensal:
        mensal, fonte_m = obter_focus_odata_mensal()
        fontes.append(fonte_m)
    if not anual:
        anual, fonte_a = obter_focus_odata_anual()
        fontes.append(fonte_a)
    return mensal, anual, " ; ".join(fontes)


def taxa_mensal_por_focus(mes: str, focus_mensal: Dict[str, Decimal], focus_anual: Dict[int, Decimal]) -> Tuple[Decimal, str]:
    if mes in focus_mensal:
        return focus_mensal[mes], "Focus/BCB mensal"
    ano = int(mes[:4])
    if ano in focus_anual:
        taxa = (Decimal("1") + focus_anual[ano]) ** (Decimal("1") / Decimal("12")) - Decimal("1")
        return taxa, "Focus/BCB anual convertido para taxa mensal equivalente"
    return IPCA_PROJECAO_MENSAL_PADRAO, "fallback local padrao"

def preencher_indices_futuros(indices: Dict[str, Decimal]) -> Tuple[Dict[str, Decimal], Dict[str, str]]:
    """Completa meses ainda nao divulgados usando Focus/BCB.

    Regra:
    1) Mantem numero-indice oficial do SIDRA/IBGE quando existir.
    2) Para meses futuros, usa expectativa Focus mensal do BCB quando disponivel.
    3) Se nao houver mensal, usa expectativa Focus anual convertida para taxa mensal equivalente.
    4) Se tudo falhar, usa IPCA_PROJECAO_MENSAL_PADRAO.
    """
    out = dict(indices)
    fontes = {m: "SIDRA/IBGE ou fallback oficial" for m in out}
    focus_mensal, focus_anual, fonte_focus = obter_focus_ipca()

    ultimo_mes = max(out)
    ultimo_indice = out[ultimo_mes]
    mes_atual = add_months(date(int(ultimo_mes[:4]), int(ultimo_mes[5:7]), 1), 1)
    fim = mes_str(add_months(max(aniversario(d) for d, _ in CRONOGRAMA), -1))

    while mes_str(mes_atual) <= fim:
        m = mes_str(mes_atual)
        # Se houver valor oficial/conhecido no fallback local, ele prevalece sobre Focus.
        if m in IPCA_FALLBACK_VARIACAO_MENSAL:
            taxa = IPCA_FALLBACK_VARIACAO_MENSAL[m]
            fonte_taxa = "fallback oficial/conhecido local"
        else:
            taxa, fonte_taxa = taxa_mensal_por_focus(m, focus_mensal, focus_anual)
        ultimo_indice = ultimo_indice * (Decimal("1") + taxa)
        out[m] = ultimo_indice
        fontes[m] = f"{fonte_taxa} | {fonte_focus}"
        mes_atual = add_months(mes_atual, 1)
    return out, fontes

def fator_ipca(indices: Dict[str, Decimal], data_aniv: date) -> Tuple[Decimal, str, str, Decimal, Decimal]:
    """Fator IPCA mensal correto para a Data de Aniversario.

    Para o aniversario do mes M, usa NIk=M-1 e NIk-1=M-2.
    """
    mes_nik = mes_str(add_months(data_aniv, -1))
    mes_nik_1 = mes_str(add_months(data_aniv, -2))
    if mes_nik not in indices or mes_nik_1 not in indices:
        raise RuntimeError(f"IPCA necessario nao disponivel: NIk={mes_nik}, NIk_1={mes_nik_1}.")
    ni_k = indices[mes_nik]
    ni_k_1 = indices[mes_nik_1]
    fator = trunc_dec(ni_k / ni_k_1, 8)
    return fator, mes_nik, mes_nik_1, ni_k, ni_k_1


def fator_ipca_prorata(indices: Dict[str, Decimal], data_aniv: date, inicio: date) -> Tuple[Decimal, str, str, Decimal, Decimal, int, int]:
    """Fator IPCA pro rata usado somente no primeiro periodo, da integralizacao ao primeiro aniversario."""
    fator_cheio, mes_nik, mes_nik_1, ni_k, ni_k_1 = fator_ipca(indices, data_aniv)
    prev_m = add_months(data_aniv, -1)
    inicio_aniv = date(prev_m.year, prev_m.month, 15)
    dup = dias_uteis(inicio, data_aniv)
    dut = dias_uteis(inicio_aniv, data_aniv)
    bruto = Decimal(str(float(ni_k / ni_k_1) ** (dup / dut)))
    return trunc_dec(bruto, 8), mes_nik, mes_nik_1, ni_k, ni_k_1, dup, dut


def fator_juros_252(du: int) -> Decimal:
    bruto = Decimal(str((1.0 + float(TAXA_AA)) ** (du / 252.0)))
    return bruto.quantize(Decimal("0.000000001"), rounding=ROUND_HALF_UP)



def proxima_data_aniversario(dt: date) -> date:
    """Proxima data de aniversario (dia 15) em ou apos dt."""
    if dt.day <= 15:
        return date(dt.year, dt.month, 15)
    prox = add_months(date(dt.year, dt.month, 1), 1)
    return date(prox.year, prox.month, 15)


def aplicar_ipca_ate(
    saldo_pu: Decimal,
    data_ipca_atual: date,
    data_aniv_alvo: date,
    indices: Dict[str, Decimal],
    fonte_mes: Dict[str, str],
) -> Tuple[Decimal, List[Dict[str, object]], date]:
    """Aplica a atualizacao monetaria mensal ate a data de aniversario alvo.

    data_ipca_atual e a ultima data ate a qual o saldo ja esta atualizado.
    Para o primeiro trecho, quando a data inicial nao e aniversario, usa pro rata por dias uteis.
    Depois, aplica os fatores mensais cheios de aniversario a aniversario.
    """
    detalhes: List[Dict[str, object]] = []
    if data_aniv_alvo < data_ipca_atual:
        raise RuntimeError("data_aniv_alvo anterior a data_ipca_atual")
    prox = proxima_data_aniversario(data_ipca_atual)
    atual = data_ipca_atual
    saldo = trunc_dec(saldo_pu, 8)

    while prox <= data_aniv_alvo:
        if atual == prox:
            # Ja esta atualizado ate esta data.
            prox_m = add_months(prox, 1)
            prox = date(prox_m.year, prox_m.month, 15)
            continue

        if atual.day == 15 and atual < prox:
            fator_c, mes_nik, mes_nik_1, ni_k, ni_k_1 = fator_ipca(indices, prox)
            prev_m = add_months(prox, -1)
            inicio_aniv = date(prev_m.year, prev_m.month, 15)
            dup_ipca = dias_uteis(inicio_aniv, prox)
            dut_ipca = dup_ipca
        else:
            fator_c, mes_nik, mes_nik_1, ni_k, ni_k_1, dup_ipca, dut_ipca = fator_ipca_prorata(indices, prox, atual)

        saldo_ini = saldo
        saldo = trunc_dec(saldo * fator_c, 8)
        detalhes.append({
            "Data_Aniv_IPCA": prox.strftime("%d/%m/%Y"),
            "Mes_NIk": mes_nik,
            "Mes_NIk_1": mes_nik_1,
            "NIk": ni_k,
            "NIk_1": ni_k_1,
            "Fonte_NIk": fonte_mes.get(mes_nik, ""),
            "DUP_IPCA": dup_ipca,
            "DUT_IPCA": dut_ipca,
            "Fator_C_IPCA": fator_c,
            "PU_Antes_IPCA": saldo_ini,
            "PU_Apos_IPCA": saldo,
        })
        atual = prox
        prox_m = add_months(prox, 1)
        prox = date(prox_m.year, prox_m.month, 15)

    return saldo, detalhes, data_aniv_alvo


def calcular_fluxo() -> Tuple[List[Dict[str, object]], str]:
    indices, fonte = obter_ipca_numero_indice_sidra()
    indices, fonte_mes = preencher_indices_futuros(indices)

    saldo_pu = trunc_dec(PU_INICIAL, 8)
    data_ref_juros = DATA_INICIO_RENTABILIDADE
    data_ipca_atual = DATA_INICIO_RENTABILIDADE
    linhas: List[Dict[str, object]] = []

    # Eventos especiais: juros dos dois primeiros periodos incorporados ao VNA.
    for i, data_incorp in enumerate(DATAS_INCORPORACAO_JUROS, start=1):
        saldo_pu, detalhes_ipca, data_ipca_atual = aplicar_ipca_ate(
            saldo_pu,
            data_ipca_atual,
            data_incorp,
            indices,
            fonte_mes,
        )
        du_incorp = dias_uteis(data_ref_juros, data_incorp)
        fj_incorp = fator_juros_252(du_incorp)
        pu_juros_incorp = trunc_dec(saldo_pu * (fj_incorp - Decimal("1")), 8)
        saldo_antes_incorp = saldo_pu
        saldo_pu = trunc_dec(saldo_pu + pu_juros_incorp, 8)
        linhas.append({
            "Evento": f"Incorporacao Juros {i}",
            "Data_Ref": data_incorp.strftime("%d/%m/%Y"),
            "Data_Pgto": "",
            "DU_Juros": du_incorp,
            "Qtde_IPCA_Aplicados": len(detalhes_ipca),
            "Ultimo_Mes_NIk": detalhes_ipca[-1]["Mes_NIk"] if detalhes_ipca else "",
            "Ultimo_Mes_NIk_1": detalhes_ipca[-1]["Mes_NIk_1"] if detalhes_ipca else "",
            "Ultimo_Fator_C_IPCA": detalhes_ipca[-1]["Fator_C_IPCA"] if detalhes_ipca else "",
            "Fator_Juros": fj_incorp,
            "TAI_Amort": Decimal("0"),
            "PU_VNa_Ini": trunc_dec(PU_INICIAL, 8) if i == 1 else linhas[-1]["PU_VNa_Fim"],
            "PU_VNa_Atualizado": saldo_antes_incorp,
            "PU_Juros": pu_juros_incorp,
            "PU_Juros_Incorporado": pu_juros_incorp,
            "PU_Amort": Decimal("0"),
            "PU_Total_Pago": Decimal("0"),
            "PU_VNa_Fim": saldo_pu,
            "Juros_R$": Decimal("0.00"),
            "Amort_R$": Decimal("0.00"),
            "PMT_Total": Decimal("0.00"),
            "Saldo_Devedor_R$": round_dec(saldo_pu * QUANTIDADE, 2),
        })
        data_ref_juros = data_incorp

    for data_pagto, perc_amort in CRONOGRAMA:
        data_aniv = aniversario(data_pagto)
        pu_vna_ini = trunc_dec(saldo_pu, 8)
        saldo_pu, detalhes_ipca, data_ipca_atual = aplicar_ipca_ate(
            saldo_pu,
            data_ipca_atual,
            data_aniv,
            indices,
            fonte_mes,
        )

        du = dias_uteis(data_ref_juros, data_pagto)
        fj = fator_juros_252(du)

        pu_vna_atualizado = trunc_dec(saldo_pu, 8)
        pu_juros = trunc_dec(pu_vna_atualizado * (fj - Decimal("1")), 8)
        pu_amort = trunc_dec(pu_vna_atualizado * perc_amort, 8)
        pu_vna_fim = trunc_dec(pu_vna_atualizado - pu_amort, 8)

        juros_rs = round_dec(pu_juros * QUANTIDADE, 2)
        amort_rs = round_dec(pu_amort * QUANTIDADE, 2)
        pmt_rs = round_dec(juros_rs + amort_rs, 2)
        saldo_rs = round_dec(pu_vna_fim * QUANTIDADE, 2)

        linhas.append({
            "Evento": "Pagamento",
            "Data_Ref": data_aniv.strftime("%d/%m/%Y"),
            "Data_Pgto": data_pagto.strftime("%d/%m/%Y"),
            "DU_Juros": du,
            "Qtde_IPCA_Aplicados": len(detalhes_ipca),
            "Ultimo_Mes_NIk": detalhes_ipca[-1]["Mes_NIk"] if detalhes_ipca else "",
            "Ultimo_Mes_NIk_1": detalhes_ipca[-1]["Mes_NIk_1"] if detalhes_ipca else "",
            "Ultimo_Fator_C_IPCA": detalhes_ipca[-1]["Fator_C_IPCA"] if detalhes_ipca else "",
            "Fator_Juros": fj,
            "TAI_Amort": perc_amort,
            "PU_VNa_Ini": pu_vna_ini,
            "PU_VNa_Atualizado": pu_vna_atualizado,
            "PU_Juros": pu_juros,
            "PU_Juros_Incorporado": Decimal("0"),
            "PU_Amort": pu_amort,
            "PU_Total_Pago": trunc_dec(pu_juros + pu_amort, 8),
            "PU_VNa_Fim": pu_vna_fim,
            "Juros_R$": juros_rs,
            "Amort_R$": amort_rs,
            "PMT_Total": pmt_rs,
            "Saldo_Devedor_R$": saldo_rs,
        })

        saldo_pu = pu_vna_fim
        data_ref_juros = data_pagto

    return linhas, fonte

def salvar_csv(linhas: List[Dict[str, object]], caminho: str) -> None:
    if not linhas:
        raise RuntimeError("Nenhuma linha calculada.")
    campos = list(linhas[0].keys())
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=campos, delimiter=";")
        w.writeheader()
        for row in linhas:
            out: Dict[str, object] = {}
            for k, v in row.items():
                out[k] = str(v).replace(".", ",") if isinstance(v, Decimal) else v
            w.writerow(out)




def _decimal_to_float(v: object) -> object:
    if isinstance(v, Decimal):
        return float(v)
    return v


def salvar_xlsx(linhas: List[Dict[str, object]], caminho: str, fonte_ipca: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.comments import Comment
    except Exception as exc:
        raise RuntimeError("Para gerar XLSX, instale openpyxl: pip install openpyxl") from exc

    if not linhas:
        raise RuntimeError("Nenhuma linha calculada.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Controle_Divida"
    ws.freeze_panes = "A2"

    campos = list(linhas[0].keys())
    ws.append(campos)
    for row in linhas:
        ws.append([_decimal_to_float(row.get(c, "")) for c in campos])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    date_cols = {"Data_Ref", "Data_Pgto"}
    money_cols = {"Juros_R$", "Amort_R$", "PMT_Total", "Saldo_Devedor_R$"}
    pu_cols = {"PU_VNa_Ini", "PU_VNa_Atualizado", "PU_Juros", "PU_Juros_Incorporado", "PU_Amort", "PU_Total_Pago", "PU_VNa_Fim"}
    factor_cols = {"Ultimo_Fator_C_IPCA", "Fator_Juros"}
    pct_cols = {"TAI_Amort"}

    for idx, nome in enumerate(campos, start=1):
        col = get_column_letter(idx)
        if nome in money_cols:
            number_format = '#,##0.00'
        elif nome in pu_cols or nome in factor_cols:
            number_format = '0.00000000'
        elif nome in pct_cols:
            number_format = '0.0000%'
        elif nome == "DU_Juros" or nome == "Qtde_IPCA_Aplicados":
            number_format = '0'
        else:
            number_format = None
        if number_format:
            for cell in ws[col][1:]:
                cell.number_format = number_format

    widths = {
        "A": 24, "B": 12, "C": 12, "D": 10, "E": 12, "F": 14, "G": 16, "H": 16,
        "I": 15, "J": 12, "K": 15, "L": 18, "M": 14, "N": 19, "O": 14, "P": 16,
        "Q": 15, "R": 15, "S": 15, "T": 15, "U": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 28

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="TabelaControleAXS09", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    for cell in ws[1]:
        if cell.value in ("Ultimo_Mes_NIk", "Ultimo_Mes_NIk_1", "Ultimo_Fator_C_IPCA"):
            cell.comment = Comment("Para eventos semestrais, mostra o ultimo mes de IPCA aplicado no bloco ate a data de aniversario.", "ChatGPT")
        if cell.value == "PU_Juros_Incorporado":
            cell.comment = Comment("Juros dos dois primeiros periodos incorporados ao VNA em 15/03/2025 e 15/09/2025, conforme escritura e 1o aditivo.", "ChatGPT")

    ws2 = wb.create_sheet("Parametros")
    params = [
        ["Campo", "Valor"],
        ["Emissao", "AXS Energia Unidade 09"],
        ["Data de Emissao", DATA_EMISSAO.strftime("%d/%m/%Y")],
        ["Data de Inicio da Rentabilidade", DATA_INICIO_RENTABILIDADE.strftime("%d/%m/%Y")],
        ["Datas de Incorporacao dos Juros", ", ".join(d.strftime("%d/%m/%Y") for d in DATAS_INCORPORACAO_JUROS)],
        ["Taxa AA", float(TAXA_AA)],
        ["Quantidade", float(QUANTIDADE)],
        ["PU Inicial", float(PU_INICIAL)],
        ["Fonte IPCA", fonte_ipca],
        ["Observacao", "Calculo sem imputar valores calculados pela Vortx; sem uso de PU, VNA, juros ou PMT da Vortx como entrada."],
    ]
    for r in params:
        ws2.append(r)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 90
    for cell in ws2["B"]:
        if cell.row == 6:
            cell.number_format = '0.00%'
        elif cell.row in (7, 8):
            cell.number_format = '#,##0.00'

    wb.save(caminho)
def imprimir_linha(data_txt: str, linhas: List[Dict[str, object]]) -> None:
    linha = next((x for x in linhas if x["Data_Pgto"] == data_txt), None)
    if not linha:
        print(f"\nLinha {data_txt}: nao calculada.")
        return
    print(f"\nLinha {data_txt}:")
    campos = [
        "Evento", "Data_Ref", "Data_Pgto", "DU_Juros", "Qtde_IPCA_Aplicados",
        "Ultimo_Mes_NIk", "Ultimo_Mes_NIk_1", "Ultimo_Fator_C_IPCA",
        "Fator_Juros", "PU_VNa_Atualizado", "PU_Juros", "PU_Amort", "Juros_R$",
        "Amort_R$", "PMT_Total", "Saldo_Devedor_R$",
    ]
    for k in campos:
        print(f"{k}: {linha.get(k, '')}")


def main() -> None:
    linhas, fonte = calcular_fluxo()
    arquivo_csv = "controle_divida_axs09_v1_focus.csv"
    arquivo_xlsx = "controle_divida_axs09_v1_focus.xlsx"
    salvar_csv(linhas, arquivo_csv)
    try:
        salvar_xlsx(linhas, arquivo_xlsx, fonte)
        print(f"Arquivo XLSX gerado: {arquivo_xlsx}")
    except Exception as exc:
        print(f"Nao foi possivel gerar XLSX: {exc}")

    print("Fonte IPCA oficial/historica:", fonte)
    print(f"Arquivo CSV gerado: {arquivo_csv}")
    imprimir_linha("15/03/2026", linhas)
    imprimir_linha("15/09/2026", linhas)
    imprimir_linha("15/09/2038", linhas)


if __name__ == "__main__":
    main()
