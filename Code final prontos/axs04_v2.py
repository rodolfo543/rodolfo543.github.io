# -*- coding: utf-8 -*-
"""
AXS 05 - CRI AXS IV / Opea 139a emissao - calculo em VALOR TOTAL com IPCA e Focus/BCB.

Objetivo
- Recriar o fluxo do CRI AXS IV sem imputar PU, VNA, juros, amortizacao ou PMT da Opea/Vortx.
- Esta versao calcula primeiro o fluxo em VALOR TOTAL da operacao e usa PU apenas como informativo/auditoria.
- Usa a regra contratual do Termo de Securitizacao consolidado pelo 1o aditamento.
- Gera CSV e XLSX editavel, no mesmo espirito das planilhas AXS 07/08/09.

Principais premissas do documento
- Valor total da emissao dos CRI: R$ 144.000.000,00, conforme Termo de Securitizacao.
- PU de referencia: R$ 1.000,00; quantidade equivalente = valor total / PU.
- Data de emissao dos CRI: 15/06/2023.
- Taxa: IPCA + 11,00% a.a., base 252 dias uteis.
- Juros e amortizacao pagos mensalmente conforme Anexo I.
- Vencimento final: 15/06/2037.

Ajuste importante
- A data da primeira integralizacao efetiva foi parametrizada como 30/06/2023. O 1o aditamento de 27/06/2023 informa que ainda nao havia integralizacao ate aquela data; o historico de pagamentos da Opea indica primeiro fator pro rata compativel com 11 dias uteis ate 17/07/2023. Nao foram imputados valores de juros, amortizacao, PMT, PU ou VNA da Opea.
- O 1o aditamento ajusta NIk como o numero-indice do segundo mes imediatamente anterior a Data de Pagamento.
  Isso esta em IPCA_LAG_MESES = 2. Se alguma memoria antiga usar o exemplo original, teste IPCA_LAG_MESES = 1.

Como rodar
    python axs05_cri_axsiv_v2_total_focus_bcb_xlsx.py

Dependencias opcionais
    pip install openpyxl python-bcb

Arquivos gerados
    controle_divida_axs05_cri_axsiv_v2_total.csv
    controle_divida_axs05_cri_axsiv_v2_total.xlsx
"""

from __future__ import annotations

import csv
import json
import ssl
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_DOWN, ROUND_HALF_UP, getcontext
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen
import importlib.util

getcontext().prec = 34

DATA_EMISSAO = date(2023, 6, 15)
DATA_INICIO_RENTABILIDADE = date(2023, 6, 30)
DATA_VENCIMENTO = date(2037, 6, 15)
PU_INICIAL = Decimal("1000.00000000")
VALOR_TOTAL_INICIAL = Decimal("144000000.00000000")
QUANTIDADE_EQUIVALENTE = VALOR_TOTAL_INICIAL / PU_INICIAL
TAXA_AA = Decimal("0.1100")
IPCA_LAG_MESES = 2
DUT_PRIMEIRO_PERIODO = 22

# Criterio ANBIMA/Termo para CRI IPCA:
# - VNE informado com 8 casas;
# - VNA truncado em 8 casas;
# - fator pro rata do IPCA truncado em 8 casas;
# - fator acumulado das variacoes de inflacao truncado em 8 casas;
# - amortizacao e fluxo financeiro truncados em 8 casas e valor financeiro final em 2 casas.
# O primeiro periodo usa dut = 22 dias uteis, conforme excecao contratual.

CRONOGRAMA_RAW = [
    ("2023-07-17", "0.1000"),
    ("2023-08-15", "0.1001"),
    ("2023-09-15", "0.1002"),
    ("2023-10-16", "0.1003"),
    ("2023-11-16", "0.1004"),
    ("2023-12-15", "0.1005"),
    ("2024-01-15", "0.1006"),
    ("2024-02-15", "0.1007"),
    ("2024-03-15", "0.1008"),
    ("2024-04-15", "0.1009"),
    ("2024-05-15", "0.1010"),
    ("2024-06-17", "0.1011"),
    ("2024-07-15", "0.1012"),
    ("2024-08-15", "0.1013"),
    ("2024-09-16", "0.1014"),
    ("2024-10-15", "0.1015"),
    ("2024-11-18", "0.1016"),
    ("2024-12-16", "0.1017"),
    ("2025-01-15", "0.2037"),
    ("2025-02-17", "0.2041"),
    ("2025-03-17", "0.2045"),
    ("2025-04-15", "0.2049"),
    ("2025-05-15", "0.2053"),
    ("2025-06-16", "0.4630"),
    ("2025-07-15", "0.4651"),
    ("2025-08-15", "0.5192"),
    ("2025-09-15", "0.5219"),
    ("2025-10-15", "0.5247"),
    ("2025-11-17", "0.5274"),
    ("2025-12-15", "0.5302"),
    ("2026-01-15", "0.4797"),
    ("2026-02-18", "0.4821"),
    ("2026-03-16", "0.4844"),
    ("2026-04-15", "0.4867"),
    ("2026-05-15", "0.4891"),
    ("2026-06-15", "0.4915"),
    ("2026-07-15", "0.4940"),
    ("2026-08-17", "0.4964"),
    ("2026-09-15", "0.4989"),
    ("2026-10-15", "0.5014"),
    ("2026-11-16", "0.5039"),
    ("2026-12-15", "0.5065"),
    ("2027-01-15", "0.5090"),
    ("2027-02-15", "0.5117"),
    ("2027-03-15", "0.5143"),
    ("2027-04-15", "0.5169"),
    ("2027-05-17", "0.5196"),
    ("2027-06-15", "0.5223"),
    ("2027-07-15", "0.5834"),
    ("2027-08-16", "0.5869"),
    ("2027-09-15", "0.5903"),
    ("2027-10-15", "0.5938"),
    ("2027-11-16", "0.5974"),
    ("2027-12-15", "0.6010"),
    ("2028-01-17", "0.6046"),
    ("2028-02-15", "0.6083"),
    ("2028-03-15", "0.6120"),
    ("2028-04-17", "0.6158"),
    ("2028-05-15", "0.6196"),
    ("2028-06-16", "0.6234"),
    ("2028-07-17", "0.6274"),
    ("2028-08-15", "0.6313"),
    ("2028-09-15", "0.6353"),
    ("2028-10-16", "0.6394"),
    ("2028-11-16", "0.6435"),
    ("2028-12-15", "0.7124"),
    ("2029-01-15", "0.7175"),
    ("2029-02-15", "0.7227"),
    ("2029-03-15", "0.7280"),
    ("2029-04-16", "0.7333"),
    ("2029-05-15", "0.7388"),
    ("2029-06-15", "0.8119"),
    ("2029-07-16", "0.8186"),
    ("2029-08-15", "0.8253"),
    ("2029-09-17", "0.8322"),
    ("2029-10-15", "0.8392"),
    ("2029-11-16", "0.8463"),
    ("2029-12-17", "0.9957"),
    ("2030-01-15", "1.0057"),
    ("2030-02-15", "0.8104"),
    ("2030-03-15", "0.8170"),
    ("2030-04-15", "0.8237"),
    ("2030-05-15", "0.9049"),
    ("2030-06-17", "0.9132"),
    ("2030-07-15", "0.9216"),
    ("2030-08-15", "1.1595"),
    ("2030-09-16", "1.1731"),
    ("2030-10-15", "1.1871"),
    ("2030-11-18", "1.2013"),
    ("2030-12-16", "1.2159"),
    ("2031-01-15", "0.9063"),
    ("2031-02-17", "0.9146"),
    ("2031-03-17", "0.9230"),
    ("2031-04-15", "0.9316"),
    ("2031-05-15", "0.9404"),
    ("2031-06-16", "1.0343"),
    ("2031-07-15", "0.9592"),
    ("2031-08-15", "1.2287"),
    ("2031-09-15", "1.2440"),
    ("2031-10-15", "1.2596"),
    ("2031-11-17", "1.2757"),
    ("2031-12-15", "1.2922"),
    ("2032-01-15", "1.3091"),
    ("2032-02-16", "1.1392"),
    ("2032-03-15", "1.1523"),
    ("2032-04-15", "1.1658"),
    ("2032-05-17", "1.1795"),
    ("2032-06-15", "1.2917"),
    ("2032-07-15", "1.2092"),
    ("2032-08-16", "1.5258"),
    ("2032-09-15", "1.5495"),
    ("2032-10-15", "1.5738"),
    ("2032-11-16", "1.5990"),
    ("2032-12-15", "1.6250"),
    ("2033-01-17", "1.6518"),
    ("2033-02-15", "1.4581"),
    ("2033-03-15", "1.4797"),
    ("2033-04-18", "1.5019"),
    ("2033-05-16", "1.5248"),
    ("2033-06-15", "1.6660"),
    ("2033-07-15", "1.6942"),
    ("2033-08-15", "1.7234"),
    ("2033-09-15", "1.7537"),
    ("2033-10-17", "1.7850"),
    ("2033-11-16", "1.8174"),
    ("2033-12-15", "1.8510"),
    ("2034-01-16", "1.8860"),
    ("2034-02-15", "1.9222"),
    ("2034-03-15", "1.9599"),
    ("2034-04-17", "1.9991"),
    ("2034-05-15", "2.0398"),
    ("2034-06-15", "2.4988"),
    ("2034-07-17", "2.5628"),
    ("2034-08-15", "2.6302"),
    ("2034-09-15", "2.7013"),
    ("2034-10-16", "2.7763"),
    ("2034-11-16", "2.8555"),
    ("2034-12-15", "2.9395"),
    ("2035-01-15", "3.0285"),
    ("2035-02-15", "3.1231"),
    ("2035-03-15", "3.2238"),
    ("2035-04-16", "3.3312"),
    ("2035-05-15", "3.4459"),
    ("2035-06-15", "3.7789"),
    ("2035-07-16", "3.9273"),
    ("2035-08-15", "4.0878"),
    ("2035-09-17", "4.2620"),
    ("2035-10-15", "4.4518"),
    ("2035-11-16", "4.6592"),
    ("2035-12-17", "4.8869"),
    ("2036-01-15", "5.1380"),
    ("2036-02-15", "6.0181"),
    ("2036-03-17", "6.4034"),
    ("2036-04-15", "6.8415"),
    ("2036-05-15", "7.3439"),
    ("2036-06-16", "7.9260"),
    ("2036-07-15", "8.6083"),
    ("2036-08-15", "9.4192"),
    ("2036-09-15", "10.3986"),
    ("2036-10-15", "11.6054"),
    ("2036-11-17", "13.1291"),
    ("2036-12-15", "15.1134"),
    ("2037-01-15", "17.8042"),
    ("2037-02-18", "21.6606"),
    ("2037-03-16", "27.6498"),
    ("2037-04-15", "38.2166"),
    ("2037-05-15", "61.8557"),
    ("2037-06-15", "100.0000"),
]

CRONOGRAMA = [(datetime.strptime(d, "%Y-%m-%d").date(), Decimal(p) / Decimal("100")) for d, p in CRONOGRAMA_RAW]

IPCA_PROJECAO_MENSAL_PADRAO = Decimal("0.0045")


def trunc_dec(x: Decimal, casas: int = 8) -> Decimal:
    return x.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_DOWN)


def round_dec(x: Decimal, casas: int = 2) -> Decimal:
    return x.quantize(Decimal("1").scaleb(-casas), rounding=ROUND_HALF_UP)


def add_months(dt: date, months: int) -> date:
    y = dt.year + (dt.month - 1 + months) // 12
    m = (dt.month - 1 + months) % 12 + 1
    return date(y, m, 1)


def mes_str(dt: date) -> str:
    return f"{dt.year:04d}-{dt.month:02d}"


def easter_date(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def feriados_nacionais(start_year: int = 2022, end_year: int = 2037) -> set[date]:
    fs: set[date] = set()
    for y in range(start_year, end_year + 1):
        pascoa = easter_date(y)
        fs.update({
            date(y, 1, 1),
            pascoa - timedelta(days=48),
            pascoa - timedelta(days=47),
            pascoa - timedelta(days=2),
            date(y, 4, 21),
            date(y, 5, 1),
            pascoa + timedelta(days=60),
            date(y, 9, 7),
            date(y, 10, 12),
            date(y, 11, 2),
            date(y, 11, 15),
            date(y, 11, 20),
            date(y, 12, 25),
        })
    return fs


FERIADOS = feriados_nacionais()


def eh_dia_util(dt: date) -> bool:
    return dt.weekday() < 5 and dt not in FERIADOS


def dias_uteis(inicio: date, fim: date) -> int:
    """Conta dias uteis em [inicio, fim)."""
    count = 0
    dt = inicio
    while dt < fim:
        if eh_dia_util(dt):
            count += 1
        dt += timedelta(days=1)
    return count


def ultimo_mes_necessario() -> str:
    max_data = max(d for d, _ in CRONOGRAMA)
    mes = add_months(date(max_data.year, max_data.month, 1), -IPCA_LAG_MESES)
    return mes_str(mes)


def obter_json_url(url: str, timeout: int = 25) -> object:
    ctx = ssl.create_default_context()
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=timeout, context=ctx) as resp:
        return json.loads(resp.read().decode("utf-8"))


def obter_ipca_numero_indice_sidra() -> Tuple[Dict[str, Decimal], str]:
    """Busca o numero-indice IPCA no SIDRA/IBGE. Se falhar, retorna vazio e o script usa Focus/projecao."""
    fim = ultimo_mes_necessario().replace("-", "")
    url = f"https://apisidra.ibge.gov.br/values/t/1737/n1/all/v/2266/p/202201-{fim}?formato=json"
    try:
        dados = obter_json_url(url, timeout=30)
        if not isinstance(dados, list) or len(dados) <= 1:
            raise RuntimeError("SIDRA sem dados")
        indices: Dict[str, Decimal] = {}
        for item in dados[1:]:
            periodo = str(item.get("D3C", ""))
            valor = str(item.get("V", "")).replace(",", ".")
            if len(periodo) == 6 and valor not in ("", "...", "-"):
                indices[f"{periodo[:4]}-{periodo[4:]}"] = Decimal(valor)
        if not indices:
            raise RuntimeError("SIDRA retornou vazio ou layout inesperado")
        return indices, f"SIDRA/IBGE Tabela 1737 v/2266 | {url}"
    except Exception as exc:
        return {}, f"SIDRA/IBGE indisponivel: {exc}"


def decimal_ptbr(valor: object) -> Decimal | None:
    if valor is None:
        return None
    txt = str(valor).strip().replace("%", "").replace(" ", "")
    if not txt or txt in ("...", "-", "--"):
        return None
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    else:
        txt = txt.replace(",", ".")
    try:
        return Decimal(txt)
    except Exception:
        return None


def parse_mes_referencia(valor: object) -> str | None:
    if valor is None:
        return None
    txt = str(valor).strip()
    if len(txt) >= 7 and txt[4] == "-" and txt[:4].isdigit() and txt[5:7].isdigit():
        return txt[:7]
    if "/" in txt:
        partes = txt.split("/")
        if len(partes) >= 2 and partes[0].isdigit() and partes[1].isdigit():
            mes = int(partes[0])
            ano = int(partes[1])
            if ano < 100:
                ano += 2000
            if 1 <= mes <= 12:
                return f"{ano:04d}-{mes:02d}"
    if len(txt) == 6 and txt.isdigit():
        return f"{txt[:4]}-{txt[4:]}"
    return None


def build_odata_url(recurso: str, params: Dict[str, str], usar_parenteses: bool = True) -> str:
    base = "https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata"
    sufixo = "()" if usar_parenteses and not recurso.endswith(")") else ""
    return f"{base}/{recurso}{sufixo}?{urlencode(params, quote_via=quote)}"


def obter_focus_odata_mensal() -> Tuple[Dict[str, Decimal], str]:
    recursos = ["ExpectativaMercadoMensais", "ExpectativasMercadoMensais"]
    params = {
        "$top": "10000",
        "$format": "json",
        "$select": "Indicador,Data,DataReferencia,Mediana",
        "$filter": "Indicador eq 'IPCA'",
        "$orderby": "Data desc",
    }
    erros: List[str] = []
    for recurso in recursos:
        for usar_parenteses in (True, False):
            try:
                dados = obter_json_url(build_odata_url(recurso, params, usar_parenteses), timeout=30)
                itens = dados.get("value", []) if isinstance(dados, dict) else []
                out: Dict[str, Tuple[str, Decimal]] = {}
                for item in itens:
                    if str(item.get("Indicador", "")).upper() != "IPCA":
                        continue
                    mes = parse_mes_referencia(item.get("DataReferencia"))
                    med = decimal_ptbr(item.get("Mediana"))
                    if not mes or med is None:
                        continue
                    taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                    data_pub = str(item.get("Data", ""))
                    if mes not in out or data_pub > out[mes][0]:
                        out[mes] = (data_pub, taxa)
                if out:
                    return {m: v[1] for m, v in out.items()}, f"Focus/BCB mensal via OData {recurso}"
            except Exception as exc:
                erros.append(f"{recurso}: {exc}")
    return {}, "Focus mensal OData indisponivel: " + " | ".join(erros[:4])


def obter_focus_odata_anual() -> Tuple[Dict[int, Decimal], str]:
    recursos = ["ExpectativasMercadoAnuais", "ExpectativaMercadoAnuais"]
    params = {
        "$top": "10000",
        "$format": "json",
        "$select": "Indicador,Data,DataReferencia,Mediana",
        "$filter": "Indicador eq 'IPCA'",
        "$orderby": "Data desc",
    }
    erros: List[str] = []
    for recurso in recursos:
        for usar_parenteses in (True, False):
            try:
                dados = obter_json_url(build_odata_url(recurso, params, usar_parenteses), timeout=30)
                itens = dados.get("value", []) if isinstance(dados, dict) else []
                out: Dict[int, Tuple[str, Decimal]] = {}
                for item in itens:
                    if str(item.get("Indicador", "")).upper() != "IPCA":
                        continue
                    ref = str(item.get("DataReferencia", "")).strip()
                    med = decimal_ptbr(item.get("Mediana"))
                    if not ref.isdigit() or med is None:
                        continue
                    ano = int(ref)
                    taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                    data_pub = str(item.get("Data", ""))
                    if ano not in out or data_pub > out[ano][0]:
                        out[ano] = (data_pub, taxa)
                if out:
                    return {a: v[1] for a, v in out.items()}, f"Focus/BCB anual via OData {recurso}"
            except Exception as exc:
                erros.append(f"{recurso}: {exc}")
    return {}, "Focus anual OData indisponivel: " + " | ".join(erros[:4])


def obter_focus_python_bcb() -> Tuple[Dict[str, Decimal], Dict[int, Decimal], str]:
    try:
        if importlib.util.find_spec("bcb") is None:
            return {}, {}, "python-bcb nao instalado"
        from bcb import Expectativas  # type: ignore
        em = Expectativas()
    except Exception as exc:
        return {}, {}, f"python-bcb indisponivel: {exc}"
    mensal: Dict[str, Decimal] = {}
    anual: Dict[int, Decimal] = {}
    mensagens: List[str] = []
    try:
        ep = em.get_endpoint("ExpectativaMercadoMensais")
        df = ep.query().filter(ep.Indicador == "IPCA").select(ep.Indicador, ep.Data, ep.DataReferencia, ep.Mediana).orderby(ep.Data.desc()).limit(20000).collect()
        temp: Dict[str, Tuple[str, Decimal]] = {}
        for item in df.to_dict("records"):
            mes = parse_mes_referencia(item.get("DataReferencia"))
            med = decimal_ptbr(item.get("Mediana"))
            if mes and med is not None:
                taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                data_pub = str(item.get("Data", ""))
                if mes not in temp or data_pub > temp[mes][0]:
                    temp[mes] = (data_pub, taxa)
        mensal = {m: v[1] for m, v in temp.items()}
        mensagens.append(f"python-bcb mensal: {len(mensal)} meses")
    except Exception as exc:
        mensagens.append(f"python-bcb mensal falhou: {exc}")
    try:
        ep = em.get_endpoint("ExpectativasMercadoAnuais")
        df = ep.query().filter(ep.Indicador == "IPCA").select(ep.Indicador, ep.Data, ep.DataReferencia, ep.Mediana).orderby(ep.Data.desc()).limit(20000).collect()
        temp2: Dict[int, Tuple[str, Decimal]] = {}
        for item in df.to_dict("records"):
            ref = str(item.get("DataReferencia", "")).strip()
            med = decimal_ptbr(item.get("Mediana"))
            if ref.isdigit() and med is not None:
                ano = int(ref)
                taxa = med / Decimal("100") if abs(med) > Decimal("0.05") else med
                data_pub = str(item.get("Data", ""))
                if ano not in temp2 or data_pub > temp2[ano][0]:
                    temp2[ano] = (data_pub, taxa)
        anual = {a: v[1] for a, v in temp2.items()}
        mensagens.append(f"python-bcb anual: {len(anual)} anos")
    except Exception as exc:
        mensagens.append(f"python-bcb anual falhou: {exc}")
    return mensal, anual, " | ".join(mensagens)


def obter_focus_ipca() -> Tuple[Dict[str, Decimal], Dict[int, Decimal], str]:
    mensal, anual, fonte_py = obter_focus_python_bcb()
    fontes = [fonte_py]
    if not mensal:
        mensal, fonte_m = obter_focus_odata_mensal()
        fontes.append(fonte_m)
    if not anual:
        anual, fonte_a = obter_focus_odata_anual()
        fontes.append(fonte_a)
    return mensal, anual, " ; ".join(fontes)


def taxa_mensal_por_focus(mes: str, focus_mensal: Dict[str, Decimal], focus_anual: Dict[int, Decimal]) -> Tuple[Decimal, str]:
    if mes in focus_mensal:
        return focus_mensal[mes], "Focus/BCB mensal"
    ano = int(mes[:4])
    if ano in focus_anual:
        taxa = (Decimal("1") + focus_anual[ano]) ** (Decimal("1") / Decimal("12")) - Decimal("1")
        return taxa, "Focus/BCB anual convertido para taxa mensal equivalente"
    return IPCA_PROJECAO_MENSAL_PADRAO, "fallback local padrao"


def preencher_indices_futuros(indices: Dict[str, Decimal]) -> Tuple[Dict[str, Decimal], Dict[str, str], str]:
    out = dict(indices)
    fontes = {m: "SIDRA/IBGE" for m in out}
    foco_m, foco_a, fonte_focus = obter_focus_ipca()
    if out:
        ultimo_mes = max(out)
        ultimo_indice = out[ultimo_mes]
        mes_atual = add_months(date(int(ultimo_mes[:4]), int(ultimo_mes[5:7]), 1), 1)
    else:
        # Base sintetica apenas para permitir projecao quando SIDRA estiver indisponivel.
        ultimo_indice = Decimal("7000.0000000000")
        mes_atual = date(2022, 1, 1)
    fim = ultimo_mes_necessario()
    while mes_str(mes_atual) <= fim:
        m = mes_str(mes_atual)
        if m not in out:
            taxa, fonte_taxa = taxa_mensal_por_focus(m, foco_m, foco_a)
            ultimo_indice = ultimo_indice * (Decimal("1") + taxa)
            out[m] = ultimo_indice
            fontes[m] = f"{fonte_taxa} | {fonte_focus}"
        else:
            ultimo_indice = out[m]
        mes_atual = add_months(mes_atual, 1)
    return out, fontes, fonte_focus


def fator_ipca_periodo(indices: Dict[str, Decimal], data_pagto: date, data_anterior: date, primeiro: bool) -> Tuple[Decimal, str, str, Decimal, Decimal, int, int]:
    mes_base = date(data_pagto.year, data_pagto.month, 1)
    mes_nik = mes_str(add_months(mes_base, -IPCA_LAG_MESES))
    mes_nik_1 = mes_str(add_months(mes_base, -IPCA_LAG_MESES - 1))
    if mes_nik not in indices or mes_nik_1 not in indices:
        raise RuntimeError(f"IPCA necessario nao disponivel: NIk={mes_nik}, NIk_1={mes_nik_1}")
    ni_k = indices[mes_nik]
    ni_k_1 = indices[mes_nik_1]
    dup = dias_uteis(data_anterior, data_pagto)
    dut = DUT_PRIMEIRO_PERIODO if primeiro else dup
    if dut <= 0:
        dut = dup or 1
    fator_bruto = Decimal(str(float(ni_k / ni_k_1) ** (dup / dut)))
    return trunc_dec(fator_bruto, 8), mes_nik, mes_nik_1, ni_k, ni_k_1, dup, dut


def fator_juros_252(du: int) -> Decimal:
    bruto = Decimal(str((1.0 + float(TAXA_AA)) ** (du / 252.0)))
    return bruto.quantize(Decimal("0.000000001"), rounding=ROUND_HALF_UP)


def calcular_fluxo() -> Tuple[List[Dict[str, object]], str, str]:
    indices_sidra, fonte_sidra = obter_ipca_numero_indice_sidra()
    indices, fontes_meses, fonte_focus = preencher_indices_futuros(indices_sidra)
    linhas: List[Dict[str, object]] = []
    saldo_total = VALOR_TOTAL_INICIAL
    data_anterior = DATA_INICIO_RENTABILIDADE
    for i, (data_pgto, perc_amort) in enumerate(CRONOGRAMA, start=1):
        primeiro = i == 1
        fator_c, mes_nik, mes_nik_1, ni_k, ni_k_1, du_ipca, dut_ipca = fator_ipca_periodo(indices, data_pgto, data_anterior, primeiro)
        total_vna_ini = saldo_total
        total_vna_atualizado = trunc_dec(total_vna_ini * fator_c, 8)
        pu_vna_ini = trunc_dec(total_vna_ini / QUANTIDADE_EQUIVALENTE, 8)
        pu_vna_atualizado = trunc_dec(total_vna_atualizado / QUANTIDADE_EQUIVALENTE, 8)
        du_juros = dias_uteis(data_anterior, data_pgto)
        fat_juros = fator_juros_252(du_juros)
        total_juros = trunc_dec(total_vna_atualizado * (fat_juros - Decimal("1")), 8)
        total_amort = trunc_dec(total_vna_atualizado * perc_amort, 8)
        total_pago = trunc_dec(total_juros + total_amort, 8)
        saldo_final_total = trunc_dec(total_vna_atualizado - total_amort, 8)
        pu_juros = trunc_dec(total_juros / QUANTIDADE_EQUIVALENTE, 8)
        pu_amort = trunc_dec(total_amort / QUANTIDADE_EQUIVALENTE, 8)
        pu_total = trunc_dec(total_pago / QUANTIDADE_EQUIVALENTE, 8)
        saldo_final_pu = trunc_dec(saldo_final_total / QUANTIDADE_EQUIVALENTE, 8)
        linhas.append({
            "Evento": i,
            "Tipo": "Pagamento",
            "Data_Ref": data_pgto,
            "Data_Pgto": data_pgto,
            "DU_Juros": du_juros,
            "DU_IPCA": du_ipca,
            "DUT_IPCA": dut_ipca,
            "Mes_NIk": mes_nik,
            "Mes_NIk_1": mes_nik_1,
            "NIk": ni_k,
            "NIk_1": ni_k_1,
            "Fonte_IPCA": fontes_meses.get(mes_nik, ""),
            "Fator_C_IPCA": fator_c,
            "Fator_Juros": fat_juros,
            "TAI_Amort": perc_amort,
            "PU_VNa_Ini": pu_vna_ini,
            "PU_VNa_Atualizado": pu_vna_atualizado,
            "PU_Juros": pu_juros,
            "PU_Amort": pu_amort,
            "PU_Total_Pago": pu_total,
            "PU_VNa_Fim": saldo_final_pu,
            "VNa_Ini_R$": round_dec(total_vna_ini, 2),
            "VNa_Atualizado_R$": round_dec(total_vna_atualizado, 2),
            "Juros_R$": round_dec(total_juros, 2),
            "Amort_R$": round_dec(total_amort, 2),
            "PMT_Total": round_dec(total_pago, 2),
            "Saldo_Devedor_R$": round_dec(saldo_final_total, 2),
        })
        saldo_total = saldo_final_total
        data_anterior = data_pgto
    return linhas, fonte_sidra, fonte_focus


def fmt_data(dt: object) -> str:
    if isinstance(dt, date):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def fmt_decimal(v: object) -> object:
    if isinstance(v, Decimal):
        return str(v).replace(".", ",")
    if isinstance(v, date):
        return fmt_data(v)
    return v


def salvar_csv(linhas: List[Dict[str, object]], caminho: Path) -> None:
    if not linhas:
        return
    with caminho.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(linhas[0].keys()), delimiter=";")
        writer.writeheader()
        for row in linhas:
            writer.writerow({k: fmt_decimal(v) for k, v in row.items()})


def salvar_xlsx(linhas: List[Dict[str, object]], caminho: Path, fonte_sidra: str, fonte_focus: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.comments import Comment
    except Exception as exc:
        print(f"[AVISO] openpyxl nao disponivel, XLSX nao gerado: {exc}")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle_Divida"
    headers = list(linhas[0].keys())
    ws.append(headers)
    for row in linhas:
        ws.append([row[h] for h in headers])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            if isinstance(cell.value, date):
                cell.number_format = "dd/mm/yyyy"
            elif isinstance(cell.value, Decimal):
                cell.value = float(cell.value)
                cell.number_format = "#,##0.00000000"
    money_cols = ["Juros_R$", "Amort_R$", "PMT_Total", "Saldo_Devedor_R$"]
    for name in money_cols:
        col = headers.index(name) + 1
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
            for c in cell:
                c.number_format = 'R$ #,##0.00'
    perc_cols = ["TAI_Amort"]
    for name in perc_cols:
        col = headers.index(name) + 1
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
            for c in cell:
                c.number_format = "0.0000%"
    tab = Table(displayName="TabelaControleAXS05CRITotal", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    for idx, h in enumerate(headers, start=1):
        width = max(12, min(35, len(h) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws["A1"].comment = Comment("AXS 05 CRI: calculo por metodologia em valor total, sem imputar PU/VNA/PMT da Opea/Vortx.", "ChatGPT")

    wp = wb.create_sheet("Parametros")
    params = [
        ("DATA_EMISSAO", DATA_EMISSAO),
        ("DATA_INICIO_RENTABILIDADE", DATA_INICIO_RENTABILIDADE),
        ("DATA_VENCIMENTO", DATA_VENCIMENTO),
        ("PU_INICIAL_REFERENCIA", PU_INICIAL),
        ("VALOR_TOTAL_INICIAL", VALOR_TOTAL_INICIAL),
        ("QUANTIDADE_EQUIVALENTE", QUANTIDADE_EQUIVALENTE),
        ("TAXA_AA", TAXA_AA),
        ("IPCA_LAG_MESES", IPCA_LAG_MESES),
        ("DUT_PRIMEIRO_PERIODO", DUT_PRIMEIRO_PERIODO),
        ("FONTE_IPCA_OFICIAL", fonte_sidra),
        ("FONTE_FOCUS", fonte_focus),
    ]
    wp.append(["Parametro", "Valor"])
    for p, v in params:
        wp.append([p, fmt_decimal(v)])
    for cell in wp[1]:
        cell.fill = header_fill
        cell.font = header_font
    wp.column_dimensions["A"].width = 32
    wp.column_dimensions["B"].width = 120
    wb.save(caminho)


def main() -> None:
    linhas, fonte_sidra, fonte_focus = calcular_fluxo()
    out_csv = Path("controle_divida_axs04_cri_axsiv_v2_total.csv")
    out_xlsx = Path("controle_divida_axs04_cri_axsiv_v2_total.xlsx")
    salvar_csv(linhas, out_csv)
    salvar_xlsx(linhas, out_xlsx, fonte_sidra, fonte_focus)
    print(f"Gerado: {out_csv.resolve()}")
    if out_xlsx.exists():
        print(f"Gerado: {out_xlsx.resolve()}")
    print(f"Fonte IPCA: {fonte_sidra}")
    print(f"Fonte Focus: {fonte_focus}")
    if linhas:
        ult = linhas[-1]
        print("Ultimo evento:", fmt_data(ult["Data_Pgto"]), "PMT", fmt_decimal(ult["PMT_Total"]), "Saldo", fmt_decimal(ult["Saldo_Devedor_R$"]))


if __name__ == "__main__":
    main()
